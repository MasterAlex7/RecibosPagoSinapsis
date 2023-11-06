import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def crearExcel(response):
    book = Workbook()
    sheet = book.active

    sheet['A1'] = 'ID Pago'
    sheet['B1'] = 'Nombre Cliente'
    sheet['C1'] = 'Numero de Contrato'
    sheet['D1'] = 'Fecha de pago'
    sheet['E1'] = 'Mensualidad'
    sheet['F1'] = 'Abono'
    sheet['G1'] = 'Descuento'

    sheet.column_dimensions['A'].width = 10
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['G'].width = 20

    sheet['A1'].font = Font(bold=True)
    sheet['B1'].font = Font(bold=True)
    sheet['C1'].font = Font(bold=True)
    sheet['D1'].font = Font(bold=True)
    sheet['E1'].font = Font(bold=True)
    sheet['F1'].font = Font(bold=True)
    sheet['G1'].font = Font(bold=True)

    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['B1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['C1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['D1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['E1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['F1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['G1'].alignment = Alignment(horizontal='center', vertical='center')

    for i in range(len(response)):
        sheet.cell(row=i+2,column=1).value = response[i]['IdPago']
        sheet.cell(row=i+2,column=2).value = response[i]['NombreCliente']
        sheet.cell(row=i+2,column=3).value = response[i]['NumeroContrato']
        sheet.cell(row=i+2,column=4).value = response[i]['FechaPago']
        sheet.cell(row=i+2,column=5).value = response[i]['Mensualidad']
        sheet.cell(row=i+2,column=6).value = response[i]['Abono']
        sheet.cell(row=i+2,column=7).value = response[i]['Descuento']

        sheet.cell(row=i+2,column=1).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=i+2,column=2).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=i+2,column=3).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=i+2,column=4).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=i+2,column=5).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=i+2,column=6).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=i+2,column=7).alignment = Alignment(horizontal='center', vertical='center')

    
    fecha = datetime.datetime.now()
    sheet.cell(row=1,column=9).value = 'Fecha de Reporte'
    sheet.cell(row=1,column=10).value = fecha.strftime("%d/%m/%Y")
    sheet.cell(row=1,column=9).font = Font(bold=True)
    sheet.cell(row=1,column=10).font = Font(bold=True)
    sheet.cell(row=1,column=9).alignment = Alignment(horizontal='center', vertical='center')
    sheet.cell(row=1,column=10).alignment = Alignment(horizontal='center', vertical='center')
    sheet.column_dimensions['J'].width = 20
    sheet.column_dimensions['I'].width = 20
    

    book.save("C:\Recibos de Pago\Reportes Mensuales\Report"+" "+fecha.strftime("%B")+fecha.strftime("%Y")+".xlsx")

if __name__ == '__main__':
    crearExcel()